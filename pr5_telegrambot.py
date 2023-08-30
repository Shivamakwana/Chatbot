import telebot
import openpyxl
import os
from telebot import types
from openpyxl.styles import Font
from datetime import datetime  # Import the datetime module

# Replace 'YOUR_TOKEN' with your actual Telegram bot token
bot = telebot.TeleBot('YOUR_TOKEN')

user_data = {}

# List of questions
questions = [
    "hi..! I am ArthTech_Bot\nI'm here for taking some Details of your Internship...\nPlease enter your Name:\n(eg. John Wick)",
    "Enter College Name:",
    "Enter Selected Course Name:",
    "Enter Starting Date:(YYYY-MM-DD)",
    "Enter Ending Date:(YYYY-MM-DD)",
    "Enter Total No. of Internship Days:",
    "Enter No. of Completed Days:",
    "Enter Total No. of Roadmaps Points:",
    "Enter No. of Completed Roadmaps Points:"
]

# list of columns
column_names = [
    "Name",
    "College Name",
    "Course Name",
    "Starting Date",
    "Ending Date",
    "Total Internship Days",
    "Completed Days",
    "Total Roadmaps Points",
    "Completed Roadmaps Points",
    "Performance"  # New column for performance
]

@bot.message_handler(commands=['start'])
def start(message):
    user_data[message.chat.id] = {'current_question': 0, 'answers': {}}
    bot.send_message(message.chat.id, questions[0])

@bot.message_handler(func=lambda message: 'current_question' in user_data.get(message.chat.id, {}))
def handle_input(message):
    current_question = user_data[message.chat.id]['current_question']
    question = questions[current_question]
    
    if question not in user_data[message.chat.id]['answers']:
        # Convert relevant inputs to appropriate format
        if current_question in [3, 4]:
            try:
                date_obj = datetime.strptime(message.text, '%Y-%m-%d')
                user_data[message.chat.id]['answers'][question] = date_obj.strftime('%Y-%m-%d')
            except ValueError:
                bot.send_message(message.chat.id, "Invalid date format. Please use YYYY-MM-DD format.")
                return
        elif current_question in [5, 6, 7, 8]:
            try:
                input_value = int(message.text)
                if current_question == 5:  # Total Internship Days
                    starting_date = datetime.strptime(user_data[message.chat.id]['answers'][questions[3]], '%Y-%m-%d')
                    ending_date = datetime.strptime(user_data[message.chat.id]['answers'][questions[4]], '%Y-%m-%d')
                    total_days_date = (ending_date - starting_date).days + 1
                    if input_value > total_days_date:
                        raise ValueError(f"Total Internship Days cannot exceed {total_days_date}")
                elif current_question == 6:  # Completed Days
                    total_days = user_data[message.chat.id]['answers'][questions[5]]
                    if input_value > total_days:
                        raise ValueError("Completed Days cannot exceed Total Internship Days")
                elif current_question == 8:  # Completed Roadmaps Points
                    total_points = user_data[message.chat.id]['answers'][questions[7]]
                    if input_value > total_points:
                        raise ValueError("Completed Roadmaps Points cannot exceed Total Roadmaps Points")       
                user_data[message.chat.id]['answers'][question] = input_value
            except ValueError as e:
                bot.send_message(message.chat.id, str(e) + ". Please enter a valid value.")
                return 
        else:
            user_data[message.chat.id]['answers'][question] = message.text
    
    if current_question == len(questions) - 1:
        create_excel(message.chat.id)
    else:
        user_data[message.chat.id]['current_question'] += 1
        bot.send_message(message.chat.id, questions[current_question + 1])

def calculate_performance(completed_points, total_points, completed_days, total_days):
    if total_days > 0 and total_points > 0:
        performance_ratio = completed_points / total_points / (completed_days / total_days)
        
        if performance_ratio == 1:
            return 'Good'
        elif performance_ratio > 1:
            return 'Excellent'
        else:
            return 'Poor'
    else:
        return 'N/A'

def create_excel(chat_id):
    wb = openpyxl.load_workbook("pr5_Internship_data.xlsx") if "pr5_Internship_data.xlsx" in os.listdir() else openpyxl.Workbook()
    sheet = wb.active
    
    # Apply bold font to column names
    bold_font = Font(bold=True)
    for column, column_name in enumerate(column_names):
        cell = sheet.cell(row=1, column=column + 1)
        cell.value = column_name
        cell.font = bold_font
    
    new_row = []
    for question in questions:
        new_row.append(user_data[chat_id]['answers'].get(question, ""))

    starting_date = datetime.strptime(user_data[chat_id]['answers'][questions[3]], '%Y-%m-%d')
    ending_date = datetime.strptime(user_data[chat_id]['answers'][questions[4]], '%Y-%m-%d')
    total_days_date = (ending_date - starting_date).days + 1

    completed_points = int(user_data[chat_id]['answers'].get(questions[-1], 0))
    total_points = int(user_data[chat_id]['answers'].get(questions[-2], 1))
    completed_days = int(user_data[chat_id]['answers'].get(questions[-3], 1))
    total_days = int(user_data[chat_id]['answers'].get(questions[-4], 1))
    
    if completed_days <= total_days and completed_points <= total_points and total_days_date >= completed_days:
        performance = calculate_performance(completed_points, total_points, completed_days, total_days)
        new_row.append(performance)
        sheet.append(new_row)
        
        file_name = "pr5_Internship_data.xlsx"
        wb.save(file_name)
        bot.send_message(chat_id, "Data has been saved in Excel file.\nIf you want to enter details of other user then type /start or <-- click on it.\n© All rights reserved to Shivam Makwana.")
    else:
        error_message = "Constraint violation:"
        if completed_days > total_days:
            error_message += "\nCompleted Days exceeds Total Internship Days."
        if completed_points > total_points:
            error_message += "\nCompleted Roadmaps Points exceeds Total Roadmaps Points."
        if total_days < completed_days:
            error_message += "\nEnding Date - Starting Date is less than Completed Days."
        bot.send_message(chat_id, error_message + "\nData not added to Excel.\nIf you want to re-enter the details type /start or <-- click on it.")

# Start the bot
bot.polling()
